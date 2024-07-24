import shutil
import os
from openpyxl import load_workbook

def copy_file_or_directory(source_path, destination_dir):
    try:
        if os.path.isdir(source_path):
            # Recursively copy directory
            dest_path = os.path.join(destination_dir, os.path.basename(source_path))
            shutil.copytree(source_path, dest_path)
            print(f"Directory '{source_path}' copied successfully.")
        else:
            # Copy file
            dest_path = os.path.join(destination_dir, os.path.basename(source_path))
            shutil.copy2(source_path, dest_path)
            print(f"File '{source_path}' copied successfully.")
    except PermissionError:
        print(f"Permission denied: '{source_path}'. Attempting to copy contents recursively.")
        if os.path.isdir(source_path):
            for root, dirs, files in os.walk(source_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    rel_path = os.path.relpath(file_path, source_path)
                    dest_path = os.path.join(destination_dir, os.path.basename(source_path), rel_path)
                    dest_dir = os.path.dirname(dest_path)
                    if not os.path.exists(dest_dir):
                        os.makedirs(dest_dir)
                    try:
                        shutil.copy2(file_path, dest_path)
                        print(f"File '{file_path}' copied successfully.")
                    except Exception as e:
                        print(f"Error copying file '{file_path}': {e}")
    except Exception as e:
        print(f"Error copying '{source_path}': {e}")

# Prompt for destination directory
destination_dir = input("Enter the destination directory: ").strip()

# Validate destination directory
if not os.path.exists(destination_dir):
    try:
        os.makedirs(destination_dir)
    except OSError as e:
        print(f"Error creating destination directory '{destination_dir}': {e}")
        exit(1)

if not os.path.isdir(destination_dir):
    print(f"Error: '{destination_dir}' is not a valid directory.")
    exit(1)

# Load Excel workbook
excel_file = 'files.xlsx'  # Update with your Excel file path
wb = load_workbook(excel_file)
ws = wb.active

# List to store source file paths
source_files = []

# Read file paths from Excel sheet (assuming they are in column A starting from row 1)
for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
    file_path = row[0]
    if file_path:  # Check if cell is not empty
        source_files.append(file_path)

# Copy each file to the destination directory
for source_file_path in source_files:
    if os.path.exists(source_file_path):
        copy_file_or_directory(source_file_path, destination_dir)
    else:
        print(f"File or directory not found: {source_file_path}")

print("All files and directories processed.")
